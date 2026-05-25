#!/usr/bin/env python
"""Batch validate Swiss customer addresses from Excel.

Usage:
    python batch_address_validator.py input.xlsx output.xlsx

First create a sample file:
    python create_sample_data.py
    python batch_address_validator.py sample_customers.xlsx validated_customers.xlsx
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Add parent directory to path for local imports
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from postal_address import PostalAddressClient, PostalAddressError


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Batch validate Swiss customer addresses from an Excel file.",
        epilog=(
            "Example:\n"
            "  python create_sample_data.py\n"
            "  python batch_address_validator.py sample_customers.xlsx validated_customers.xlsx"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("input_file", help="Path to input Excel file (.xlsx)")
    parser.add_argument("output_file", help="Path to output Excel file (.xlsx)")
    parser.add_argument(
        "--api-key",
        dest="api_key",
        default=None,
        help="RapidAPI key (defaults to RAPIDAPI_KEY environment variable)",
    )
    return parser.parse_args()


def validate_batch(input_file: str, output_file: str, api_key: str | None = None) -> None:
    """Read addresses from Excel, validate with Swiss API, write results."""
    import os

    api_key = api_key or os.getenv("RAPIDAPI_KEY")
    if not api_key:
        print("Error: RAPIDAPI_KEY environment variable not set.")
        sys.exit(1)

    try:
        client = PostalAddressClient(country="ch", api_key=api_key)
    except PostalAddressError as exc:
        print(f"Error initializing client: {exc}")
        sys.exit(1)

    # Open input workbook
    try:
        wb_in = openpyxl.load_workbook(input_file)
        ws_in = wb_in.active
    except Exception as exc:
        print(f"Error reading input file: {exc}")
        sys.exit(1)

    # Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title

    # Copy header row and add validation columns
    headers = []
    for cell in ws_in[1]:
        if cell.value:
            headers.append(cell.value)
    headers.extend(["ValidationStatus", "NormalizedStreet", "NormalizedPostalCode", "NormalizedCity", "ErrorMessage"])
    ws_out.append(headers)

    # Color codes for output
    valid_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    invalid_fill = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")  # Light red
    error_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold

    processed = 0
    valid_count = 0

    # Process each row
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=False), start=2):
        original_values = [cell.value for cell in row]
        if not any(original_values):  # Skip empty rows
            continue

        # Extract address components
        street = None
        house_number = None
        postal_code = None
        city = None

        # Try to find columns by header name
        for col_idx, cell in enumerate(row):
            header = ws_in.cell(row=1, column=col_idx + 1).value
            if header and header.lower() in ("street", "strasse", "rue", "via"):
                street = cell.value
            elif header and header.lower() in ("housenumber", "hnr", "numero", "nr"):
                house_number = cell.value
            elif header and header.lower() in ("postalcode", "plz", "zip", "code postal"):
                postal_code = cell.value
            elif header and header.lower() in ("city", "stadt", "ville", "citta"):
                city = cell.value

        # Write original values to output
        out_row = list(original_values)

        # Validate address
        if street and postal_code and city:
            try:
                result = client.validate_address(
                    street=street,
                    house_number=house_number,
                    zip_code=postal_code,
                    city=city,
                )
                if result and result.Flag in (1, 2):
                    out_row.extend(
                        [
                            "valid",
                            result.Street or street,
                            result.Zipcode or postal_code,
                            result.City or city,
                            "",
                        ]
                    )
                    fill = valid_fill
                    valid_count += 1
                elif result:
                    out_row.extend(["invalid", street, postal_code, city, result.FlagText or "Address not valid"])
                    fill = invalid_fill
                else:
                    out_row.extend(["invalid", street, postal_code, city, "API Call failed"])
                    fill = error_fill
            except PostalAddressError as exc:
                out_row.extend(["error", street, postal_code, city, str(exc)])
                fill = error_fill
        else:
            out_row.extend(["error", street or "", postal_code or "", city or "", "Missing required fields"])
            fill = error_fill

        # Write to output and apply color
        ws_out.append(out_row)
        for col_idx, _ in enumerate(out_row, start=1):
            ws_out.cell(row=row_idx, column=col_idx).fill = fill

        processed += 1
        if processed % 10 == 0:
            print(f"Processed {processed} rows...")

    # Auto-adjust column widths
    for column in ws_out.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_out.column_dimensions[column_letter].width = adjusted_width

    # Save output
    try:
        wb_out.save(output_file)
        print(f"\nResults saved to {output_file}")
        print(f"Total processed: {processed}")
        print(f"Valid addresses: {valid_count}")
        print(f"Invalid/Error: {processed - valid_count}")
    except Exception as exc:
        print(f"Error writing output file: {exc}")
        sys.exit(1)


if __name__ == "__main__":
    args = parse_args()
    input_file = args.input_file
    output_file = args.output_file

    if not Path(input_file).exists():
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)

    validate_batch(input_file, output_file, api_key=args.api_key)
